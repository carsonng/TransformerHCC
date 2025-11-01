import os
import torch
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import numpy as np
import torch.nn.functional as F
from torch.utils.data import DataLoader
from sklearn.metrics import confusion_matrix, roc_auc_score
from readdata.readdata_internal import MedicalImageDataset
from models.LSTM import ConvLSTMNet
from tools.tools import calculate_metrics


class InternalModelEvaluator:
    def __init__(self, model, project_name, device='cuda:0'):
        self.model = model
        self.project_name = project_name
        self.device = device
        self.criterion = torch.nn.CrossEntropyLoss().to(device)

        self.patient_data = {
            'hcc': {
                'small_train': ['baozhongxic', 'cailiyuan', 'chenguanshui', 'chenjiang', 'chenjiayana',
                                'chenmingchangb',
                                'chenxikun', 'dengpeiyoub', 'fanghaizhou', 'fengzhihong', 'hanweimina', 'hanying',
                                'heruguanga', 'huangchengpeng', 'huangxianyue', 'huangyuhuab', 'jianghouming',
                                'kongzhiqin',
                                'lianganxianb', 'lianghaibo', 'liangweiqiang', 'liangxiaofeng', 'lichangqib',
                                'lihanxiong',
                                'lijianwei', 'linjinjib', 'linjinjic', 'liquanhai', 'liudayin', 'liufanlong',
                                'liuweiguang', 'liyudongb', 'lizhongmeib', 'luocunqianb', 'luoquanguoa', 'luozhizhao',
                                'luxiaoping', 'oukaihua', 'shishengmin', 'tangguohua', 'wangmingguang',
                                'wangmingguangc',
                                'wangxiaowu', 'wangxiuchun', 'weichengfang', 'wenzhennan', 'wujisheng', 'xieshenquan',
                                'xionghao', 'yangjianhongb', 'yangmulan', 'yangyongqian', 'yaojinmaob', 'yutianyong',
                                'cengguangming', 'cengweisen', 'zhengzhiyuan', 'zhoushangqing', 'zhoushidia',
                                'zouqingjiang', 'caikaizheng', 'chenlibing', 'chenyutian', 'fangdengwen', 'fushaoman',
                                'huangziying -a', 'huangziying -b'],
                'small_test': ['chendingwei', 'chenweilin', 'dengpeiyouc', 'hanweiminb', 'huangshenglong', 'hufeng',
                               'leixinnana', 'liaozhenxuan', 'likongxin', 'liqianjun', 'liuxiangwen', 'lizhongmeia',
                               'luhuanliang', 'luojian', 'raoliangping', 'tanxiuming', 'tianbing', 'wangziganga',
                               'wuxiaoming', 'yeyutai', 'zhangzhaohe', 'zhongpeixing', 'zhongshunmei', 'zhoushidib',
                               'chengweiyi', 'chenyuxianga', 'huangshima'],
                'large_train': ['anquan', 'biyan', 'caixilina', 'cenjianggu', 'chenhuiping', 'chenjiacun',
                                'chenjinshui',
                                'chenrichoua', 'chenweihong', 'chenweiminga', 'chenwen', 'chenyongsheng',
                                'chenyuetian', 'chenzhiwei', 'chenzhongxin', 'chenzibing', 'cuiyangxiu', 'denglianjun',
                                'dengpeiyou', 'dengshizhonga', 'dengwendao', 'dengyinyinga', 'fengjiwen', 'fengruiming',
                                'ganqimei', 'gefuliang', 'guziyang', 'hanpingan', 'heruguangb', 'heweidong',
                                'heyueqiang',
                                'huangcaidi', 'huangchengyue', 'huangjiajian', 'huangjianhua', 'huangjingliang',
                                'huangmiaochenga', 'huangshenghui', 'huangshudi', 'huangsonglaia', 'huangtianyanga',
                                'huangyuhua', 'huangzhicheng', 'huhaohua', 'jiangjianguo', 'jiangjiufen',
                                'lianganxiana',
                                'lianghaize', 'lianghe', 'liangsaozai', 'liangyihuana', 'lichangqia', 'lihaihuia',
                                'lihongxia', 'lijianwena', 'likangrena', 'limaoa', 'linbinglin', 'lingwencai',
                                'linxizhonga', 'liqingren', 'liruifang', 'lishixiang', 'liuheming', 'liuhongyan',
                                'liujiahe',
                                'liuqingtang', 'liuyinqiu', 'liyingguo', 'liyoumao', 'liyudonga', 'longguangxiang',
                                'luocunqian',
                                'luoliying', 'luoquanguoc', 'luyuzhena', 'maishengguang', 'qiaoshengjun', 'qinruiwen',
                                'tangruijian', 'tangyanzhena', 'tanzuyia', 'tanzuyib', 'wangenhuia', 'wuhairong',
                                'wuxiaopeng', 'wuxuan', 'wuyuanbiao', 'wuzhijuan', 'xiaojianrong', 'xiaozuokun',
                                'xuweisong',
                                'yangjianfenga', 'yangmulian', 'yangzhengjie', 'yanshulan', 'yezhuo',
                                'yuanzongwua', 'cengzhaoronga', 'zhangjinlun', 'zhangrenjun', 'zhangshenglu',
                                'zhaopeng',
                                'zhongweihong', 'zhongwen', 'zouriyue', 'yaojinmaoa', 'chenmingchang'],
                'large_test': ['baozhongxia', 'caiyuzhen', 'chenchengjiu', 'chenjiayanb', 'chenshulin', 'chenweimingb',
                               'chenyizhong', 'chenyuhui', 'daixiafeng', 'dengtailiang', 'dongruqiang', 'guomaozi',
                               'hexingliang', 'huangfuxiang', 'huangjixian', 'huangrongnan', 'huangshukui',
                               'huangxinzong', 'huangzhiwei', 'laishuangxionga', 'liaohuoyang', 'liguanzhao',
                               'likangrenb',
                               'linfeng', 'liqingquan', 'liqiubo', 'liubo', 'liuhuowen', 'liukangduo', 'liyazhen',
                               'liyuanfeia', 'lujinkuna', 'luweiquan', 'panxusheng', 'qiuyia', 'tanxiaoyun',
                               'wenruntang',
                               'wuxingwen', 'wuzhilong', 'yangjianhong', 'yanpeixing', 'yezhong', 'cengguorong',
                               'zhangxie',
                               'zhongdehuaa', 'hejinsheng']
            },
            'xgl': {
                'small_train': ['caijianwei', 'chenhaiquan', 'chenyongcheng', 'fuchuanzhen', 'ganhaoji', 'guanxunqiang',
                                'hanyan', 'hejinhua', 'hexingda', 'huangjianhuab', 'huangxiongjie', 'hufujian',
                                'huhenglong', 'huhuiyong', 'hujiancong', 'jinqinzhong', 'leiyi', 'liangzhiqin',
                                'lilijun',
                                'limaob', 'lishujuan', 'liugang', 'liujiandong', 'liweilong', 'lixiande', 'lixiaolai',
                                'lizhendang', 'luohanmin', 'luoquanguob', 'luyanlin', 'mojianzao', 'pengyongbiao',
                                'wangzigangb', 'wengeliang', 'wubinbiao', 'wujunqing', 'xiaoguohe', 'yangjingyao',
                                'yangsujie', 'yangzongtao', 'yuanhaizhang', 'yuyiheng', 'yuzhihao', 'zhanghui',
                                'zhengluhuan', 'zhengyaying', 'zhongshunmeib', 'zhurenteng'],
                'small_test': ['chennafeng', 'fangyanhong', 'gaozhiyong', 'helibo', 'huangmingjie', 'hujiancheng',
                               'lianglizhub', 'lihaihuib', 'liqiang', 'liuhuming', 'liwenzheng', 'longhuijian',
                               'moliewei',
                               'qishaolin', 'tangjun', 'wangyanna', 'wuxiang', 'yubiao', 'cengyunqiang', 'zhangyusheng',
                               'zhengronggui'],
                'large_train': ['caisonggui', 'caiweiting', 'caizhuofang', 'chenchengwenb', 'chenjiandong',
                                'chenjinshuib',
                                'chenqingquan', 'chenshidong', 'chenwanzhu', 'dengjieying', 'fanwenhua', 'fenghualang',
                                'guohanxiong', 'hedeqianga', 'heguiyin', 'hongyangde', 'huangmiaochengb',
                                'huangshaohui',
                                'huangsonglaib', 'huoxinlin', 'jiangjiasheng', 'jiangshifeng', 'liangjunb', 'liaowubin',
                                'lijian', 'lijuanjuan', 'linjinsheng', 'linjunyuan', 'liweijun', 'lixinbing',
                                'luolixia',
                                'luolousheng', 'mengtongtong', 'pengjiyong', 'qiuyib', 'ruanjiaolan', 'suchunlan',
                                'suyongshou', 'tangyanzhenb', 'taobinghua', 'wangenhuib', 'wangmingguangb', 'wangxia',
                                'wuchunlong', 'wuqinghui', 'wuzhangyue', 'xiaoshuang', 'xuqian', 'yezhixiong',
                                'yiyuechu',
                                'yuanzongwub', 'zhangdaoxian', 'zhanglei', 'zhangnengpiao'],
                'large_test': ['caixilinb', 'chengziyang', 'chenmugen', 'duzehui', 'guoxiangyong', 'huangkui',
                               'huangyunqiang', 'jianghai', 'liaoqingfeng', 'lijianpei', 'lixin', 'lizirong',
                               'machenlin',
                               'moxiuchao', 'shihesheng', 'tanghui', 'tianguirong', 'wangqiankun', 'wuhongguang',
                               'xianglifang', 'yangzhuhui', 'cengxiangxiong', 'zhangyiyue']
            },
            'dn': {
                'small_train': ['caowenkao', 'chenhuanhao', 'dengshizhongb', 'heyuanpeng', 'huangchengpengb',
                                'huangshenglongb', 'laishuangxiongb', 'liangheb', 'liaoweichao', 'lijianweib',
                                'likangrenc',
                                'linguanghui', 'liucaiyoub', 'lixianmei', 'liyuanfeib', 'luyuzhenb', 'xiedongxing',
                                'yanyaofeng', 'cengguangmingb', 'zhangzhijian'],
                'small_test': ['guchunfa', 'leixinnanb', 'lihaiquan', 'linxizhongb', 'lizhongmeic', 'pangtaihe',
                               'yangjianfeng', 'zhanggouwa'],
                'large_train': ['baozhongxib', 'chenbaoling', 'chenrichoub', 'chenshuilin', 'chenwei', 'chenyuxiangb',
                                'hedeqiangb', 'hongshuhua', 'huanghai', 'huangtianyangb', 'hujinwen', 'jiangqishui',
                                'lailiangyuan', 'lianglizhua', 'liangyihuanb', 'lihao', 'liquanhaib', 'liuweihuang',
                                'lujinkunb', 'ruanhuaman', 'wangdawen', 'xielianfeng', 'yangjianfengb', 'zhangwenbiao'],
                'large_test': ['chenliuguang', 'dengyinyingb', 'heshaoling', 'huangzhongwen', 'hujinwenb', 'laiyongxin',
                               'liucaiyou', 'tanglizhi', 'yanjin', 'zhongdehuab']
            },
            'fnh': {
                'small_train': ['chenshijie', 'hewenyu', 'huangwenhaob', 'maichuyan', 'songjiqiang', 'wangjun',
                                'wuyansen',
                                'yaohanyuan'],
                'small_test': ['gandongsheng', 'linxingliang', 'wangdongsheng', 'zhaoxueping'],
                'large_train': ['chenjiaying', 'chenyan', 'chenyongqiang', 'gaomanchan', 'gehaotian', 'hejiajun',
                                'huangjiandong', 'huangwenhaoa', 'huangxiaohong', 'kuangjiewen', 'liangruijing',
                                'linshaobo',
                                'linyawen', 'liufangfang', 'liuxueli', 'wangshujun', 'wangyunchuan', 'xuweituan',
                                'yaotongye', 'cengzhaorongb'],
                'large_test': ['caoying', 'chenshaozhong', 'fengjianwei', 'hufeipei', 'linshumei', 'wangquan',
                               'yanghuashou',
                               'zhongxiaoli']
            }
        }

        self.small_hcc = self.patient_data['hcc']['small_train'] + self.patient_data['hcc']['small_test']
        self.large_hcc = self.patient_data['hcc']['large_train'] + self.patient_data['hcc']['large_test']

        self.small_benign = []
        self.large_benign = []
        for category in ['xgl', 'dn', 'fnh']:
            for size_type in ['small_train', 'small_test']:
                self.small_benign.extend(self.patient_data[category][size_type])
            for size_type in ['large_train', 'large_test']:
                self.large_benign.extend(self.patient_data[category][size_type])

        self.small_tumor_patients = self.small_hcc + self.small_benign
        self.large_tumor_patients = self.large_hcc + self.large_benign

        self.result_dir = f'./result_internal/{project_name}'
        os.makedirs(self.result_dir, exist_ok=True)

    def evaluate_model(self, test_loader, loop):
        self.model.eval()
        results = []

        with torch.no_grad():
            test_true = []
            test_pred = []
            test_prob = []
            test_num_correct = 0

            for voxel1, voxel2, voxel3, voxel4, voxel5, label, name in test_loader:
                voxel1, voxel2, voxel3, voxel4, voxel5 = voxel1.to(self.device), voxel2.to(self.device), \
                                                         voxel3.to(self.device), voxel4.to(self.device), \
                                                         voxel5.to(self.device)
                label = label.to(self.device)

                logits = self.model(voxel1, voxel2, voxel3, voxel4, voxel5)
                loss = self.criterion(logits, label)

                pred = logits.argmax(dim=1)
                probabilities = F.softmax(logits, dim=1).cpu().numpy()

                for n, l, p, prob in zip(name, label.cpu().numpy(), pred.cpu().numpy(), probabilities):
                    results.append({
                        'patientname': n,
                        'pred': p,
                        'label': l,
                        'probability_0': prob[0],
                        'probability_1': prob[1],
                    })

                test_true.extend(label.cpu().numpy())
                test_pred.extend(pred.cpu().numpy())
                test_prob.extend(probabilities)
                test_num_correct += torch.eq(pred, label).sum().float().item()

            acc = test_num_correct / len(test_loader.dataset)
            sensitivity, specificity, auc_value, cm = calculate_metrics(test_true, test_pred, test_prob)

            self._save_confusion_matrix(cm, loop)

            df = pd.DataFrame(results)
            self._save_predictions_to_excel(df, loop)

            small_metrics = self._calculate_group_metrics(df, self.small_tumor_patients)
            large_metrics = self._calculate_group_metrics(df, self.large_tumor_patients)

            return acc, auc_value, sensitivity, specificity, small_metrics, large_metrics

    def _save_confusion_matrix(self, cm, loop):
        plt.figure()
        sns.heatmap(cm, annot=True, fmt='d', cmap='Blues', cbar=True)
        plt.ylabel('true label')
        plt.xlabel('predicted label')
        plt.savefig(f'{self.result_dir}/confusion_matrix_{loop}.png')
        plt.close()

    def _save_predictions_to_excel(self, df, loop):
        small_data = df[df['patientname'].isin(self.small_tumor_patients)]
        large_data = df[df['patientname'].isin(self.large_tumor_patients)]

        excel_path = f'{self.result_dir}/predictions_details_{loop}.xlsx'
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            small_data.to_excel(writer, sheet_name='small', index=False)
            large_data.to_excel(writer, sheet_name='large', index=False)

            from openpyxl.styles import PatternFill
            red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

            workbook = writer.book
            sheet1 = workbook['small']
            sheet2 = workbook['large']

            for idx, row in small_data.iterrows():
                if row['pred'] != row['label']:
                    for col in range(1, len(small_data.columns) + 1):
                        sheet1.cell(row=idx + 2, column=col).fill = red_fill

            for idx, row in large_data.iterrows():
                if row['pred'] != row['label']:
                    for col in range(1, len(large_data.columns) + 1):
                        sheet2.cell(row=idx + 2, column=col).fill = red_fill

    def _calculate_group_metrics(self, df, patient_group):
        group_data = df[df['patientname'].isin(patient_group)]
        if len(group_data) == 0:
            return 0, 0, 0, 0

        true_labels = group_data['label'].values
        pred_labels = group_data['pred'].values
        pred_probs = group_data[['probability_0', 'probability_1']].values

        cm = confusion_matrix(true_labels, pred_labels)
        sensitivity = cm[1, 1] / (cm[1, 1] + cm[1, 0]) if (cm[1, 1] + cm[1, 0]) > 0 else 0
        specificity = cm[0, 0] / (cm[0, 0] + cm[0, 1]) if (cm[0, 0] + cm[0, 1]) > 0 else 0
        acc = (cm[0, 0] + cm[1, 1]) / np.sum(cm)
        auc_value = roc_auc_score(true_labels, pred_probs[:, 1])

        return acc, auc_value, sensitivity, specificity

    def format_metrics_output(self, loop, acc, auc_value, sensitivity, specificity,
                              small_metrics, large_metrics):
        mixed_metrics = f"Mixed - acc: {acc:.4f} auc:{auc_value:.4f} " \
                        f"sensitivity:{sensitivity} specificity:{specificity}\n"

        small_acc, small_auc, small_sensitivity, small_specificity = small_metrics
        large_acc, large_auc, large_sensitivity, large_specificity = large_metrics

        small_metrics_str = f"Small Tumor - ACC: {small_acc:.4f}, AUC: {small_auc:.4f}, " \
                            f"Sensitivity: {small_sensitivity:.4f}, Specificity: {small_specificity:.4f}\n"
        large_metrics_str = f"Large Tumor - ACC: {large_acc:.4f}, AUC: {large_auc:.4f}, " \
                            f"Sensitivity: {large_sensitivity:.4f}, Specificity: {large_specificity:.4f}\n"

        print(f"Loop {loop}:")
        print(mixed_metrics)
        print(small_metrics_str)
        print(large_metrics_str)

        return mixed_metrics, small_metrics_str, large_metrics_str


def main():
    project_name = 'LSTM'
    model = ConvLSTMNet()

    evaluator = InternalModelEvaluator(model, project_name)

    metrics = {
        'mixed': {'acc': [], 'auc': [], 'sensitivity': [], 'specificity': []},
        'small': {'acc': [], 'auc': [], 'sensitivity': [], 'specificity': []},
        'large': {'acc': [], 'auc': [], 'sensitivity': [], 'specificity': []}
    }

    for loop in range(10):
        print(f"Processing loop {loop}")

        test_dataset = MedicalImageDataset(mode='test')
        test_loader = DataLoader(test_dataset, batch_size=32, shuffle=False, pin_memory=True)

        model_path = f'./weights/{project_name}/{loop}.pth'
        model.load_state_dict(torch.load(model_path))
        model = model.to(evaluator.device)

        acc, auc_value, sensitivity, specificity, small_metrics, large_metrics = \
            evaluator.evaluate_model(test_loader, loop)

        metrics['mixed']['acc'].append(acc)
        metrics['mixed']['auc'].append(auc_value)
        metrics['mixed']['sensitivity'].append(sensitivity)
        metrics['mixed']['specificity'].append(specificity)

        small_acc, small_auc, small_sensitivity, small_specificity = small_metrics
        metrics['small']['acc'].append(small_acc)
        metrics['small']['auc'].append(small_auc)
        metrics['small']['sensitivity'].append(small_sensitivity)
        metrics['small']['specificity'].append(small_specificity)

        large_acc, large_auc, large_sensitivity, large_specificity = large_metrics
        metrics['large']['acc'].append(large_acc)
        metrics['large']['auc'].append(large_auc)
        metrics['large']['sensitivity'].append(large_sensitivity)
        metrics['large']['specificity'].append(large_specificity)

        mixed_str, small_str, large_str = evaluator.format_metrics_output(
            loop, acc, auc_value, sensitivity, specificity, small_metrics, large_metrics
        )

        with open(f'{evaluator.result_dir}/{project_name}_internal_test.txt', 'a') as f:
            f.write(mixed_str)
            f.write(small_str)
            f.write(large_str)

    with open(f'{evaluator.result_dir}/{project_name}_internal_test.txt', 'a') as f:
        for group_name, group_metrics in metrics.items():
            f.write(f"{group_name} acc = ({np.mean(group_metrics['acc']):.4f} ± {np.std(group_metrics['acc']):.4f})\n")
            f.write(f"{group_name} auc = ({np.mean(group_metrics['auc']):.4f} ± {np.std(group_metrics['auc']):.4f})\n")
            f.write(
                f"{group_name} sensitivity = ({np.mean(group_metrics['sensitivity']):.4f} ± {np.std(group_metrics['sensitivity']):.4f})\n")
            f.write(
                f"{group_name} specificity = ({np.mean(group_metrics['specificity']):.4f} ± {np.std(group_metrics['specificity']):.4f})\n")


if __name__ == "__main__":
    main()



