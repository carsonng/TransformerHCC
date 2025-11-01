import os
import torch
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import numpy as np
import torch.nn.functional as F
from torch.utils.data import DataLoader
from sklearn.metrics import confusion_matrix, roc_auc_score
from readdata.readdata_external import Classification
from models.LSTM import ConvLSTMNet
from tools.tools import calculate_metrics


class ModelEvaluator:
    def __init__(self, model, project_name, device='cuda:0'):
        self.model = model
        self.project_name = project_name
        self.device = device
        self.criterion = torch.nn.CrossEntropyLoss().to(device)

        self.patient_data = {
            'small': [
                '2-2', '59', '68-2', '80', '113-2', '114-2', '127-2', '159', '183', '184-2', '223-2', '247-2',
                '272', '287', '299', '299-2', '315-2', '321', '324', 'L7-2', 'L18', 'L18-2', 'L25', 'L27',
                'L27-2', 'L42', 'L58', 'L67', 'L74-2', 'L76', 'L114-2', 'L118-2', 'L123', 'L127', 'L131',
                'L138', 'L146', 'L174-2', 'L185-2', 'L203-2', 'L220', 'L221-2', 'L235', 'L243', 'L248',
                'L253-2', 'L253-3', 'L258', 'L273', 'L293', 'L302-3', 'L331-2', 'L348-2', 'L351', 'L372',
                'L412-2', 'L426-2', 'L431', 'L435-2', 'L507', 'L544', 'L546-2', 'L547', 'L552-3', 'L578',
                'L580', 'L581'
            ],
            'small_add': [
                'huguixi', 'lianzhanglai', 'linbaoren', 'liujingfan', 'liyanhao', 'luoguangshan', 'luohongdi',
                'qianxiuzhen', 'renxuedian', 'tianguixiang', 'wangwenming', 'wangzhendan', 'weibinxian -a',
                'weibinxian -b', 'wentong', 'wukefang', 'wuxin', 'xieshaoqing', 'zhangdechao', 'zhangguoqiang',
                'zhanzhenxiong', 'zhaoaifen', 'zhengjinyang', 'zhengwang', 'zhoucanyan', 'zhouguoliang',
                'zhuoxiaowen', 'zoubaiqi'
            ],
            'large': [
                '1', '2', '15', '18', '20', '23', '25', '27', '38', '39', '66', '68', '78', '78-2', '83',
                '83-2', '89', '90', '92', '100', '106', '113', '114', '118', '119', '127', '132', '139',
                '140', '152', '163', '163-2', '177', '182', '184', '186', '187', '189', '195', '202',
                '206', '223', '247', '255', '257', '259', '263', '270', '272-2', '290', '292', '296',
                '302', '310', '315', '341', 'L7', 'L26', 'L35', 'L37', 'L41', 'L50', 'L74', 'L75', 'L85',
                'L86', 'L102', 'L114', 'L118', 'L137', 'L142', 'L147', 'L154', 'L168', 'L174', 'L185',
                'L187', 'L196', 'L203', 'L207', 'L218', 'L221', 'L224', 'L226', 'L239', 'L253', 'L257',
                'L257-2', 'L261', 'L262', 'L267', 'L283', 'L285', 'L302', 'L302-2', 'L315', 'L315-2',
                'L317', 'L324', 'L331', 'L346', 'L348', 'L364', 'L374', 'L377', 'L378', 'L385', 'L399',
                'L412', 'L426', 'L430', 'L433', 'L435', 'L443', 'L443-2', 'L444', 'L453', 'L453-2', 'L456',
                'L473', 'L473-2', 'L484', 'L485', 'L486', 'L496', 'L498', 'L501', 'L521', 'L537', 'L542',
                'L545', 'L546', 'L552', 'L552-2', 'L560', 'L563', 'L567', 'L578', 'L579', 'L582', 'L583',
                'L584', 'L585', 'L586'
            ]
        }

        self.small_tumor_patients = self.patient_data['small'] + self.patient_data['small_add']
        self.large_tumor_patients = self.patient_data['large']
        self.total_patients = self.small_tumor_patients + self.large_tumor_patients

        self.result_dir = f'./result_external/{project_name}'
        os.makedirs(self.result_dir, exist_ok=True)

    def evaluate_model(self, test_loader, loop):
        self.model.eval()
        results = []

        with torch.no_grad():
            test_true = []
            test_pred = []
            test_prob = []
            test_num_correct = 0

            for voxel1, voxel2, voxel3, voxel4, voxel5, label, name in test_loader:
                voxel1, voxel2, voxel3, voxel4, voxel5 = voxel1.to(self.device), voxel2.to(self.device), \
                                                         voxel3.to(self.device), voxel4.to(self.device), \
                                                         voxel5.to(self.device)
                label = label.to(self.device)

                logits = self.model(voxel1, voxel2, voxel3, voxel4, voxel5)
                loss = self.criterion(logits, label)

                pred = logits.argmax(dim=1)
                probabilities = F.softmax(logits, dim=1).cpu().numpy()

                for n, l, p, prob in zip(name, label.cpu().numpy(), pred.cpu().numpy(), probabilities):
                    results.append({
                        'patientname': n,
                        'pred': p,
                        'label': l,
                        'probability_0': prob[0],
                        'probability_1': prob[1],
                    })

                test_true.extend(label.cpu().numpy())
                test_pred.extend(pred.cpu().numpy())
                test_prob.extend(probabilities)
                test_num_correct += torch.eq(pred, label).sum().float().item()

            acc = test_num_correct / len(test_loader.dataset)
            sensitivity, specificity, auc_value, cm = calculate_metrics(test_true, test_pred, test_prob)

            self._save_confusion_matrix(cm, loop)

            self._save_predictions_to_excel(results, loop)

            df = pd.DataFrame(results)
            small_metrics = self._calculate_group_metrics(df, self.small_tumor_patients)
            large_metrics = self._calculate_group_metrics(df, self.large_tumor_patients)

            return acc, auc_value, sensitivity, specificity, small_metrics, large_metrics

    def _save_confusion_matrix(self, cm, loop):
        plt.figure()
        sns.heatmap(cm, annot=True, fmt='d', cmap='Blues', cbar=True)
        plt.ylabel('true label')
        plt.xlabel('predicted label')
        plt.savefig(f'{self.result_dir}/confusion_matrix_{loop}.png')
        plt.close()

    def _save_predictions_to_excel(self, results, loop):
        df = pd.DataFrame(results)
        small_data = df[df['patientname'].isin(self.small_tumor_patients)]
        large_data = df[df['patientname'].isin(self.large_tumor_patients)]

        excel_path = f'{self.result_dir}/predictions_details_{loop}.xlsx'
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            small_data.to_excel(writer, sheet_name='small', index=False)
            large_data.to_excel(writer, sheet_name='large', index=False)

            from openpyxl.styles import PatternFill
            red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

            workbook = writer.book
            sheet1 = workbook['small']
            sheet2 = workbook['large']

            for idx, row in small_data.iterrows():
                if row['pred'] != row['label']:
                    for col in range(1, len(small_data.columns) + 1):
                        sheet1.cell(row=idx + 2, column=col).fill = red_fill

            for idx, row in large_data.iterrows():
                if row['pred'] != row['label']:
                    for col in range(1, len(large_data.columns) + 1):
                        sheet2.cell(row=idx + 2, column=col).fill = red_fill

    def _calculate_group_metrics(self, df, patient_group):
        group_data = df[df['patientname'].isin(patient_group)]
        if len(group_data) == 0:
            return 0, 0, 0, 0

        true_labels = group_data['label'].values
        pred_labels = group_data['pred'].values
        pred_probs = group_data[['probability_0', 'probability_1']].values

        cm = confusion_matrix(true_labels, pred_labels)
        sensitivity = cm[1, 1] / (cm[1, 1] + cm[1, 0]) if (cm[1, 1] + cm[1, 0]) > 0 else 0
        specificity = cm[0, 0] / (cm[0, 0] + cm[0, 1]) if (cm[0, 0] + cm[0, 1]) > 0 else 0
        acc = (cm[0, 0] + cm[1, 1]) / np.sum(cm)
        auc_value = roc_auc_score(true_labels, pred_probs[:, 1])

        return acc, auc_value, sensitivity, specificity

    def format_metrics_output(self, loop, acc, auc_value, sensitivity, specificity,
                              small_metrics, large_metrics):
        mixed_metrics = f"Mixed - acc: {acc:.4f} auc:{auc_value:.4f} " \
                        f"sensitivity:{sensitivity} specificity:{specificity}\n"

        small_acc, small_auc, small_sensitivity, small_specificity = small_metrics
        large_acc, large_auc, large_sensitivity, large_specificity = large_metrics

        small_metrics_str = f"Small Tumor - ACC: {small_acc:.4f}, AUC: {small_auc:.4f}, " \
                            f"Sensitivity: {small_sensitivity:.4f}, Specificity: {small_specificity:.4f}\n"
        large_metrics_str = f"Large Tumor - ACC: {large_acc:.4f}, AUC: {large_auc:.4f}, " \
                            f"Sensitivity: {large_sensitivity:.4f}, Specificity: {large_specificity:.4f}\n"

        print(f"Loop {loop}:")
        print(mixed_metrics)
        print(small_metrics_str)
        print(large_metrics_str)

        return mixed_metrics, small_metrics_str, large_metrics_str


def main():
    test_dataset = Classification()
    test_loader = DataLoader(test_dataset, batch_size=32, shuffle=False, pin_memory=True)

    project_name = 'LSTM'
    model = LSTM()

    evaluator = ModelEvaluator(model, project_name)

    metrics = {
        'mixed': {'acc': [], 'auc': [], 'sensitivity': [], 'specificity': []},
        'small': {'acc': [], 'auc': [], 'sensitivity': [], 'specificity': []},
        'large': {'acc': [], 'auc': [], 'sensitivity': [], 'specificity': []}
    }

    for loop in range(10):
        print(f"Processing loop {loop}")

        model_path = f'./weights/{project_name}/{loop}.pth'
        model.load_state_dict(torch.load(model_path))
        model = model.to(evaluator.device)

        acc, auc_value, sensitivity, specificity, small_metrics, large_metrics = \
            evaluator.evaluate_model(test_loader, loop)

        metrics['mixed']['acc'].append(acc)
        metrics['mixed']['auc'].append(auc_value)
        metrics['mixed']['sensitivity'].append(sensitivity)
        metrics['mixed']['specificity'].append(specificity)

        small_acc, small_auc, small_sensitivity, small_specificity = small_metrics
        metrics['small']['acc'].append(small_acc)
        metrics['small']['auc'].append(small_auc)
        metrics['small']['sensitivity'].append(small_sensitivity)
        metrics['small']['specificity'].append(small_specificity)

        large_acc, large_auc, large_sensitivity, large_specificity = large_metrics
        metrics['large']['acc'].append(large_acc)
        metrics['large']['auc'].append(large_auc)
        metrics['large']['sensitivity'].append(large_sensitivity)
        metrics['large']['specificity'].append(large_specificity)

        mixed_str, small_str, large_str = evaluator.format_metrics_output(
            loop, acc, auc_value, sensitivity, specificity, small_metrics, large_metrics
        )

        with open(f'{evaluator.result_dir}/{project_name}_external_test.txt', 'a') as f:
            f.write(mixed_str)
            f.write(small_str)
            f.write(large_str)

    with open(f'{evaluator.result_dir}/{project_name}_external_test.txt', 'a') as f:
        for group_name, group_metrics in metrics.items():
            f.write(f"{group_name} acc = ({np.mean(group_metrics['acc']):.4f} ± {np.std(group_metrics['acc']):.4f})\n")
            f.write(f"{group_name} auc = ({np.mean(group_metrics['auc']):.4f} ± {np.std(group_metrics['auc']):.4f})\n")
            f.write(
                f"{group_name} sensitivity = ({np.mean(group_metrics['sensitivity']):.4f} ± {np.std(group_metrics['sensitivity']):.4f})\n")
            f.write(
                f"{group_name} specificity = ({np.mean(group_metrics['specificity']):.4f} ± {np.std(group_metrics['specificity']):.4f})\n")


if __name__ == "__main__":
    main()



